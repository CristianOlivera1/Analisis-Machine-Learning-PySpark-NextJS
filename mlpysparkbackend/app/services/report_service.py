"""
Report Service - Business logic for generating reports and exporting data
Handles report generation in JSON, Excel, and CSV formats
"""

import os
import logging
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.storage import DatasetStorage, ModelStorage
from app.utils.exceptions import (
    ValidationError,
    NotFoundError,
    BadRequestError,
    InternalServerError
)

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating reports and exporting data"""
    
    # Excel styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=14)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def generate_json_report(model_id: Optional[str] = None) -> Dict[str, Any]:
        """
        Generate a JSON report for a specific model or all models
        
        Args:
            model_id: ID of the model to generate report for (None for all models)
            
        Returns:
            Dictionary containing the report data
            
        Raises:
            NotFoundError: If model_id is provided but model not found
        """
        try:
            if model_id:
                # Single model report
                model = ModelStorage.get(model_id)
                if not model:
                    raise NotFoundError("Modelo", model_id)
                
                model_info = ModelStorage.get_info(model_id)
                dataset_info = DatasetStorage.get_info(model_info['dataset_id'])
                
                report = {
                    'report_type': 'single_model',
                    'generated_at': datetime.now().isoformat(),
                    'model': {
                        'id': model_info['id'],
                        'type': model_info['model_type'],
                        'algorithm': model_info['algorithm'],
                        'created_at': model_info['created_at'],
                        'features': model_info['features'],
                        'target': model_info.get('target'),
                        'parameters': model_info['params'],
                        'metrics': model_info['metrics'],
                        'dataset': {
                            'id': dataset_info['id'] if dataset_info else None,
                            'filename': dataset_info['filename'] if dataset_info else None
                        },
                        'data_split': {
                            'train_size': model_info['train_size'],
                            'test_size': model_info['test_size']
                        }
                    }
                }
                
                logger.info(f"JSON report generated for model: {model_id}")
                
            else:
                # All models report
                models = ModelStorage.list_all()
                
                report = {
                    'report_type': 'all_models',
                    'generated_at': datetime.now().isoformat(),
                    'summary': {
                        'total_models': len(models),
                        'by_type': ReportService._count_by_field(models, 'model_type'),
                        'by_algorithm': ReportService._count_by_field(models, 'algorithm')
                    },
                    'models': []
                }
                
                for model_info in models:
                    dataset_info = DatasetStorage.get_info(model_info['dataset_id'])
                    
                    report['models'].append({
                        'id': model_info['id'],
                        'type': model_info['model_type'],
                        'algorithm': model_info['algorithm'],
                        'created_at': model_info['created_at'],
                        'metrics': model_info['metrics'],
                        'dataset': {
                            'id': dataset_info['id'] if dataset_info else None,
                            'filename': dataset_info['filename'] if dataset_info else None
                        }
                    })
                
                logger.info(f"JSON report generated for all models ({len(models)} models)")
            
            return report
            
        except NotFoundError:
            raise
        except Exception as e:
            logger.error(f"Error generating JSON report: {str(e)}")
            raise InternalServerError(f"Error al generar reporte JSON: {str(e)}")
    
    @staticmethod
    def generate_excel_report(
        model_id: Optional[str] = None,
        include_data: bool = False
    ) -> BytesIO:
        """
        Generate an Excel report with model metrics and information
        
        Args:
            model_id: ID of the model (None for all models comparison)
            include_data: Whether to include sample data in the report
            
        Returns:
            BytesIO buffer containing the Excel file
            
        Raises:
            NotFoundError: If model_id is provided but model not found
        """
        try:
            buffer = BytesIO()
            
            if model_id:
                # Single model detailed report
                model = ModelStorage.get(model_id)
                if not model:
                    raise NotFoundError("Modelo", model_id)
                
                ReportService._create_single_model_excel(model, buffer, include_data)
                logger.info(f"Excel report generated for model: {model_id}")
                
            else:
                # Multiple models comparison report
                models = ModelStorage.list_all()
                if not models:
                    raise BadRequestError("No hay modelos disponibles para generar reporte")
                
                ReportService._create_comparison_excel(models, buffer)
                logger.info(f"Excel comparison report generated for {len(models)} models")
            
            buffer.seek(0)
            return buffer
            
        except (NotFoundError, BadRequestError):
            raise
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise InternalServerError(f"Error al generar reporte Excel: {str(e)}")
    
    # ==================== Private Helper Methods ====================
    
    @staticmethod
    def _create_single_model_excel(
        model: Dict[str, Any],
        buffer: BytesIO,
        include_data: bool
    ) -> None:
        """Create detailed Excel report for a single model"""
        
        model_info = {k: v for k, v in model.items() if k != 'pipeline_model'}
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            workbook = writer.book
            
            # 1. Summary Sheet
            ReportService._create_summary_sheet(workbook, model_info)
            
            # 2. Metrics Sheet
            ReportService._create_metrics_sheet(workbook, model_info)
            
            # 3. Parameters Sheet
            ReportService._create_parameters_sheet(workbook, model_info)
            
            # 4. Confusion Matrix Sheet (if classification)
            if model_info['model_type'] == 'classification':
                ReportService._create_confusion_matrix_sheet(workbook, model_info)
            
            # 5. Sample Data Sheet (if requested)
            if include_data:
                ReportService._create_sample_data_sheet(
                    workbook,
                    model_info['dataset_id']
                )
    
    @staticmethod
    def _create_comparison_excel(models: List[Dict], buffer: BytesIO) -> None:
        """Create Excel report comparing multiple models"""
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            workbook = writer.book
            
            # 1. Overview Sheet
            ReportService._create_overview_sheet(workbook, models)
            
            # 2. Metrics Comparison Sheet
            ReportService._create_metrics_comparison_sheet(workbook, models)
            
            # 3. Details by Type
            for model_type in ['classification', 'regression', 'clustering']:
                type_models = [m for m in models if m['model_type'] == model_type]
                if type_models:
                    ReportService._create_type_comparison_sheet(
                        workbook,
                        type_models,
                        model_type
                    )
    
    @staticmethod
    def _create_summary_sheet(workbook: Workbook, model_info: Dict) -> None:
        """Create summary information sheet"""
        
        ws = workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Model Report Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Model Information
        row = 3
        info_data = [
            ('Model ID', model_info['id']),
            ('Model Type', model_info['model_type'].upper()),
            ('Algorithm', model_info['algorithm']),
            ('Created At', model_info['created_at']),
            ('Dataset ID', model_info['dataset_id']),
            ('Target Variable', model_info.get('target', 'N/A')),
            ('Number of Features', len(model_info['features'])),
            ('Train Size', model_info['train_size']),
            ('Test Size', model_info['test_size'])
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Features List
        ws[f'A{row + 1}'] = "Features:"
        ws[f'A{row + 1}'].font = Font(bold=True)
        row += 2
        
        for feature in model_info['features']:
            ws[f'A{row}'] = f"• {feature}"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    @staticmethod
    def _create_metrics_sheet(workbook: Workbook, model_info: Dict) -> None:
        """Create metrics sheet"""
        
        ws = workbook.create_sheet("Metrics")
        
        # Title
        ws['A1'] = "Model Performance Metrics"
        ws['A1'].font = ReportService.TITLE_FONT
        ws.merge_cells('A1:B1')
        
        # Metrics
        row = 3
        ws['A3'] = "Metric"
        ws['B3'] = "Value"
        ws['A3'].font = ReportService.HEADER_FONT
        ws['B3'].font = ReportService.HEADER_FONT
        ws['A3'].fill = ReportService.HEADER_FILL
        ws['B3'].fill = ReportService.HEADER_FILL
        
        row = 4
        metrics = model_info['metrics']
        
        for metric_name, metric_value in metrics.items():
            if not isinstance(metric_value, (list, dict)):
                ws[f'A{row}'] = metric_name.replace('_', ' ').title()
                
                # Format value
                if isinstance(metric_value, float):
                    ws[f'B{row}'] = round(metric_value, 4)
                else:
                    ws[f'B{row}'] = metric_value
                
                row += 1
        
        # Apply borders
        for r in range(3, row):
            ws[f'A{r}'].border = ReportService.BORDER
            ws[f'B{r}'].border = ReportService.BORDER
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def _create_parameters_sheet(workbook: Workbook, model_info: Dict) -> None:
        """Create model parameters sheet"""
        
        ws = workbook.create_sheet("Parameters")
        
        # Title
        ws['A1'] = "Model Parameters"
        ws['A1'].font = ReportService.TITLE_FONT
        ws.merge_cells('A1:B1')
        
        # Headers
        ws['A3'] = "Parameter"
        ws['B3'] = "Value"
        ws['A3'].font = ReportService.HEADER_FONT
        ws['B3'].font = ReportService.HEADER_FONT
        ws['A3'].fill = ReportService.HEADER_FILL
        ws['B3'].fill = ReportService.HEADER_FILL
        
        # Parameters
        row = 4
        params = model_info['params']
        
        for param_name, param_value in params.items():
            ws[f'A{row}'] = param_name
            ws[f'B{row}'] = str(param_value)
            ws[f'A{row}'].border = ReportService.BORDER
            ws[f'B{row}'].border = ReportService.BORDER
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
    
    @staticmethod
    def _create_confusion_matrix_sheet(workbook: Workbook, model_info: Dict) -> None:
        """Create confusion matrix sheet for classification models"""
        
        metrics = model_info['metrics']
        if 'confusion_matrix' not in metrics:
            return
        
        ws = workbook.create_sheet("Confusion Matrix")
        
        # Title
        ws['A1'] = "Confusion Matrix"
        ws['A1'].font = ReportService.TITLE_FONT
        
        confusion_matrix = metrics['confusion_matrix']
        
        if isinstance(confusion_matrix, list) and len(confusion_matrix) > 0:
            # Create DataFrame
            df = pd.DataFrame(
                confusion_matrix,
                columns=[f"Predicted {i}" for i in range(len(confusion_matrix[0]))],
                index=[f"Actual {i}" for i in range(len(confusion_matrix))]
            )
            
            # Write to sheet starting at row 3
            row = 3
            
            # Headers
            ws[f'A{row}'] = ""
            for col_idx, col_name in enumerate(df.columns, start=2):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = col_name
                cell.font = ReportService.HEADER_FONT
                cell.fill = ReportService.HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for idx, row_name in enumerate(df.index, start=row + 1):
                ws[f'A{idx}'] = row_name
                ws[f'A{idx}'].font = Font(bold=True)
                
                for col_idx, value in enumerate(df.iloc[idx - row - 1], start=2):
                    cell = ws.cell(row=idx, column=col_idx)
                    cell.value = int(value)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = ReportService.BORDER
            
            # Adjust column widths
            for col in range(1, len(df.columns) + 2):
                ws.column_dimensions[chr(64 + col)].width = 15
    
    @staticmethod
    def _create_sample_data_sheet(workbook: Workbook, dataset_id: str) -> None:
        """Create sheet with sample data from the dataset"""
        
        dataset = DatasetStorage.get(dataset_id)
        if not dataset:
            return
        
        df_spark = dataset['dataframe']
        
        # Get sample (first 100 rows)
        df_pandas = df_spark.limit(100).toPandas()
        
        ws = workbook.create_sheet("Sample Data")
        
        # Title
        ws['A1'] = "Sample Data (First 100 rows)"
        ws['A1'].font = ReportService.TITLE_FONT
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df_pandas, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                
                # Header styling
                if r_idx == 3:
                    cell.font = ReportService.HEADER_FONT
                    cell.fill = ReportService.HEADER_FILL
                    cell.alignment = Alignment(horizontal='center')
                
                cell.border = ReportService.BORDER
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_overview_sheet(workbook: Workbook, models: List[Dict]) -> None:
        """Create overview sheet for multiple models"""
        
        ws = workbook.create_sheet("Overview", 0)
        
        # Title
        ws['A1'] = "Models Comparison Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Summary statistics
        row = 3
        ws[f'A{row}'] = "Total Models:"
        ws[f'B{row}'] = len(models)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Generated At:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        
        # Count by type
        ws[f'A{row}'] = "Models by Type:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        type_counts = ReportService._count_by_field(models, 'model_type')
        for model_type, count in type_counts.items():
            ws[f'A{row}'] = f"  {model_type.title()}"
            ws[f'B{row}'] = count
            row += 1
        
        row += 1
        
        # Count by algorithm
        ws[f'A{row}'] = "Models by Algorithm:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        algo_counts = ReportService._count_by_field(models, 'algorithm')
        for algorithm, count in algo_counts.items():
            ws[f'A{row}'] = f"  {algorithm}"
            ws[f'B{row}'] = count
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    @staticmethod
    def _create_metrics_comparison_sheet(
        workbook: Workbook,
        models: List[Dict]
    ) -> None:
        """Create sheet comparing metrics across all models"""
        
        ws = workbook.create_sheet("Metrics Comparison")
        
        # Title
        ws['A1'] = "Models Metrics Comparison"
        ws['A1'].font = ReportService.TITLE_FONT
        
        # Prepare data
        data = []
        for model in models:
            row_data = {
                'Model ID': model['id'][:8] + '...',  # Shortened ID
                'Type': model['model_type'],
                'Algorithm': model['algorithm'],
                'Created': model['created_at'][:10]  # Date only
            }
            
            # Add key metrics based on type
            metrics = model['metrics']
            if model['model_type'] == 'classification':
                row_data['Accuracy'] = metrics.get('accuracy', 'N/A')
                row_data['F1 Score'] = metrics.get('f1_score', 'N/A')
            elif model['model_type'] == 'regression':
                row_data['RMSE'] = metrics.get('rmse', 'N/A')
                row_data['R²'] = metrics.get('r2', 'N/A')
            elif model['model_type'] == 'clustering':
                row_data['Silhouette'] = metrics.get('silhouette', 'N/A')
            
            data.append(row_data)
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(data)
        
        # Write headers
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=c_idx)
            cell.value = col_name
            cell.font = ReportService.HEADER_FONT
            cell.fill = ReportService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = ReportService.BORDER
        
        # Write data
        for r_idx, row in enumerate(df.itertuples(index=False), start=4):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                
                if isinstance(value, float):
                    cell.value = round(value, 4)
                else:
                    cell.value = value
                
                cell.border = ReportService.BORDER
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_type_comparison_sheet(
        workbook: Workbook,
        models: List[Dict],
        model_type: str
    ) -> None:
        """Create detailed comparison sheet for specific model type"""
        
        sheet_name = f"{model_type.title()} Models"
        ws = workbook.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = f"{model_type.title()} Models Detailed Comparison"
        ws['A1'].font = ReportService.TITLE_FONT
        
        # Collect all metrics for this type
        all_metrics = set()
        for model in models:
            all_metrics.update(model['metrics'].keys())
        
        # Remove complex metrics
        all_metrics = {m for m in all_metrics if not isinstance(
            models[0]['metrics'].get(m), (list, dict)
        )}
        
        # Prepare data
        data = []
        for model in models:
            row_data = {
                'Model ID': model['id'][:12] + '...',
                'Algorithm': model['algorithm']
            }
            
            for metric in sorted(all_metrics):
                value = model['metrics'].get(metric, 'N/A')
                if isinstance(value, float):
                    row_data[metric.replace('_', ' ').title()] = round(value, 4)
                elif not isinstance(value, (list, dict)):
                    row_data[metric.replace('_', ' ').title()] = value
            
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Write to sheet with formatting
        ReportService._write_dataframe_to_sheet(ws, df, start_row=3)
    
    @staticmethod
    def _write_dataframe_to_sheet(
        ws,
        df: pd.DataFrame,
        start_row: int = 1
    ) -> None:
        """Write pandas DataFrame to Excel sheet with formatting"""
        
        # Write headers
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=c_idx)
            cell.value = col_name
            cell.font = ReportService.HEADER_FONT
            cell.fill = ReportService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = ReportService.BORDER
        
        # Write data
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = ReportService.BORDER
                
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        # Auto-adjust column widths
        for idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = chr(64 + idx)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _apply_excel_formatting(ws, df: pd.DataFrame) -> None:
        """Apply professional formatting to Excel worksheet"""
        
        # Format headers
        for cell in ws[1]:
            cell.font = ReportService.HEADER_FONT
            cell.fill = ReportService.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ReportService.BORDER
        
        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.border = ReportService.BORDER
                cell.alignment = Alignment(vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _count_by_field(items: List[Dict], field: str) -> Dict[str, int]:
        """Count items grouped by a specific field"""
        
        counts = {}
        for item in items:
            value = item.get(field, 'Unknown')
            counts[value] = counts.get(value, 0) + 1
        return counts
